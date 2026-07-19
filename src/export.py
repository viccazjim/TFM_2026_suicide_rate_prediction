"""
Builds the flat, denormalised tables the Power BI dashboard consumes,
and writes them to a single formatted .xlsx workbook.

This is presentation-layer packaging, not new analysis: every table
here is assembled from artifacts other pipeline stages already produce
(df_development, the persisted CatBoost/SARIMAX predictions, the
persisted production model + scaler, and the clustering functions in
src/clustering.py) — nothing in this module fits a model or computes a
result the rest of the pipeline doesn't already stand behind. Keeping
it that way means the dashboard can never show a number the thesis
itself doesn't.
"""

import numpy as np
import pandas as pd
from typing import cast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_suicide_rate_panel_table(
    df: pd.DataFrame,
    predictor_features: list[str],
    cluster_lookup: pd.DataFrame,
    target: str,
    id_col: str = "Code",
) -> pd.DataFrame:
    """
    Wide country-year panel for the map and trend pages: target,
    every determinant, and each country's Region/Cluster label —
    everything Power BI needs for a single visual without a lookup.

    Parameters
    ----------
    df : pd.DataFrame
        df_development (or equivalent country-year panel).
    predictor_features : list[str]
        Determinant columns to include (e.g. build_predictor_list()'s
        output).
    cluster_lookup : pd.DataFrame
        Must contain `id_col`, "Cluster", "Region" — one row per
        country, e.g. build_cluster_lookup_table()'s output.
    target : str
        Target column name (e.g. "Suicide rate").
    id_col : str, default "Code"

    Returns
    -------
    pd.DataFrame
        One row per country-year.
    """
    cols = [id_col, "Country", "Year", target] + list(predictor_features)
    panel = df[cols].merge(
        cluster_lookup[[id_col, "Cluster", "Region"]], on=id_col, how="left"
    )
    return panel


def build_predictions_table(
    catboost_predictions: pd.DataFrame,
    sarimax_predictions: pd.DataFrame,
    region_lookup: pd.DataFrame,
    id_col: str = "Code",
) -> pd.DataFrame:
    """
    Long-format CatBoost vs SARIMAX predictions for the predictions
    page — one row per (country, year, model), so a single Power BI
    visual can filter/compare both models with a legend rather than
    needing two side-by-side columns.

    Parameters
    ----------
    catboost_predictions : pd.DataFrame
        predict.py's output (outputs/tables/predictions.parquet) —
        must contain `id_col`, "Country", "Year", "Predicted suicide rate".
    sarimax_predictions : pd.DataFrame
        06_visualize_predictions.py's SARIMAX forecast output
        (outputs/tables/predictions_temporal.parquet) — same columns.
    region_lookup : pd.DataFrame
        Must contain `id_col`, "Region".
    id_col : str, default "Code"

    Returns
    -------
    pd.DataFrame
        Columns: [id_col, "Country", "Year", "Predicted_Rate", "Model", "Region"].
    """
    cat = cast(
        pd.DataFrame, catboost_predictions[[id_col, "Country", "Year", "Predicted suicide rate"]]
    ).copy()
    cat["Model"] = "CatBoost"
    sar = cast(
        pd.DataFrame, sarimax_predictions[[id_col, "Country", "Year", "Predicted suicide rate"]]
    ).copy()
    sar["Model"] = "SARIMAX +1 exog"

    combined = pd.concat([cat, sar], ignore_index=True)
    combined = combined.rename(columns={"Predicted suicide rate": "Predicted_Rate"})
    combined = combined.merge(region_lookup[[id_col, "Region"]], on=id_col, how="left")
    return combined


def build_trend_with_predictions_table(
    df_history: pd.DataFrame,
    predictions_table: pd.DataFrame,
    target: str,
    id_col: str = "Code",
) -> pd.DataFrame:
    """
    One long table combining the historical trend with every
    prediction model's forecast, for a single Power BI line chart
    (X=Year, Y=Rate, Legend=Series) that reproduces
    plot_predictions_trend()'s "actual flowing into predicted" look —
    Power BI has no equivalent of that function's connector-point
    trick built in, so it's replicated here in the data itself: the
    last historical year is duplicated as the first point of every
    model's series, so each dashed forecast line starts exactly where
    the solid "Actual" line ends, instead of floating with a gap.

    Parameters
    ----------
    df_history : pd.DataFrame
        Must contain `id_col`, "Country", "Year", `target` — typically
        df_development.
    predictions_table : pd.DataFrame
        build_predictions_table()'s output — must contain `id_col`,
        "Country", "Year", "Predicted_Rate", "Model".
    target : str
        Target column name in df_history (e.g. "Suicide rate").
    id_col : str, default "Code"

    Returns
    -------
    pd.DataFrame
        Columns: [id_col, "Country", "Year", "Rate", "Series"]. One
        "Actual" series per country, plus one series per country per
        model in `predictions_table`, each starting with a duplicated
        connector row.
    """
    actual = cast(
        pd.DataFrame, df_history[[id_col, "Country", "Year", target]]
    ).rename(columns={target: "Rate"})
    actual["Series"] = "Actual"

    connector_frames = [actual]
    for model_name, model_preds in predictions_table.groupby("Model"):
        for code, country_preds in model_preds.groupby(id_col):
            country_name = country_preds["Country"].iloc[0]
            hist_country = actual.loc[actual[id_col] == code].sort_values("Year")
            if hist_country.empty:
                continue
            last_year = hist_country["Year"].iloc[-1]
            last_rate = hist_country["Rate"].iloc[-1]

            connector_row = pd.DataFrame(
                {
                    id_col: [code],
                    "Country": [country_name],
                    "Year": [last_year],
                    "Rate": [last_rate],
                    "Series": [model_name],
                }
            )
            forecast_rows = cast(
                pd.DataFrame,
                country_preds.sort_values("Year")[[id_col, "Country", "Year", "Predicted_Rate"]],
            ).rename(columns={"Predicted_Rate": "Rate"})
            forecast_rows["Series"] = model_name

            connector_frames.append(pd.concat([connector_row, forecast_rows], ignore_index=True))

    return pd.concat(connector_frames, ignore_index=True)


def build_shap_importance_table(shap_values, feature_names: list[str]) -> pd.DataFrame:
    """
    Mean |SHAP value| per feature, ranked descending — for the
    determinants page's importance chart.

    Parameters
    ----------
    shap_values : shap.Explanation
        Output of src.explainability.compute_shap_values().
    feature_names : list[str]
        Column order matching shap_values' own column order (the same
        predictor list passed to compute_shap_values()).

    Returns
    -------
    pd.DataFrame
        Columns: ["Feature", "Mean_Abs_SHAP"], sorted descending.
    """
    mean_abs_shap = np.abs(shap_values.values).mean(axis=0)
    table = pd.DataFrame({"Feature": feature_names, "Mean_Abs_SHAP": mean_abs_shap})
    return table.sort_values("Mean_Abs_SHAP", ascending=False).reset_index(drop=True)


def build_model_comparison_table(result_tables: dict) -> pd.DataFrame:
    """
    Stacks the four panel-model result tables (Option A/B x Test/Val)
    into one long table tagged by Option and Split, for the model
    comparison page.

    Parameters
    ----------
    result_tables : dict[tuple[str, str], pd.DataFrame]
        Keyed by (option_label, split_label), e.g.
        {("Option A", "Test"): table_A_test, ...} — each value must
        contain "Model", "CV RMSE", "RMSE", "MAE", "R²", "Time (s)"
        (build_results_table()'s output format).

    Returns
    -------
    pd.DataFrame
        Columns: ["Model", "Option", "Split", "CV RMSE", "RMSE", "MAE", "R²", "Time (s)"].
    """
    frames = []
    for (option, split), table in result_tables.items():
        t = table.copy()
        t["Option"] = option
        t["Split"] = split
        frames.append(t)
    combined = pd.concat(frames, ignore_index=True)
    return cast(
        pd.DataFrame,
        combined[["Model", "Option", "Split", "CV RMSE", "RMSE", "MAE", "R²", "Time (s)"]],
    )


def build_cluster_lookup_table(
    country_codes,
    cluster_labels,
    pca_coords: np.ndarray,
    region_map: dict,
) -> pd.DataFrame:
    """
    One row per country: cluster assignment, a priori region, and PCA
    coordinates — feeds both build_suicide_rate_panel_table()'s
    Region/Cluster columns and the clusters page's scatter plot
    directly.

    Parameters
    ----------
    country_codes : array-like
        ISO codes, one per country, same order as cluster_labels/pca_coords.
    cluster_labels : array-like
        Output of src.clustering.run_kmeans()'s first return value.
    pca_coords : np.ndarray, shape (n_countries, 2)
        Output of src.clustering.compute_pca_coords()'s first return value.
    region_map : dict[str, str]
        e.g. src.config.EU_REGIONS.

    Returns
    -------
    pd.DataFrame
        Columns: ["Code", "Cluster", "PCA_Component_1", "PCA_Component_2", "Region"].
    """
    table = pd.DataFrame(
        {
            "Code": list(country_codes),
            "Cluster": [f"Cluster {c}" for c in cluster_labels],
            "PCA_Component_1": pca_coords[:, 0],
            "PCA_Component_2": pca_coords[:, 1],
        }
    )
    # .map(region_map) is standard, documented pandas usage (Series.map
    # accepts a dict for value substitution) — the installed stub's
    # overloads for .map() only cover a callable, not a Mapping, so this
    # is a stub gap, not a real type error (same as 06_visualize_predictions.py's
    # EU_REGIONS mapping).
    table["Region"] = table["Code"].map(region_map)  # type: ignore[arg-type]
    return table


def write_powerbi_workbook(tables: dict, path: str) -> str:
    """
    Writes every table to its own formatted sheet in a single .xlsx —
    bold white-on-navy headers, frozen header row, autofilter, and a
    column width fit to content, so the file is usable straight from
    Power BI's "Get Data > Excel" without manual cleanup.

    Parameters
    ----------
    tables : dict[str, pd.DataFrame]
        {sheet_name: table}. Sheet names should already respect Excel's
        31-character limit — not truncated here, so a name that's too
        long raises when openpyxl tries to create the sheet.
    path : str
        Output .xlsx path.

    Returns
    -------
    str
        `path`, unchanged — returned for convenient chaining/logging.
    """
    wb = Workbook()
    # A freshly-created Workbook() always has exactly one default sheet,
    # so .active is never actually None here — the assert gives pyright
    # the same guarantee it can't infer from the (correctly) Optional
    # stub type on its own, with a real runtime check behind it.
    assert wb.active is not None
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2C5F7C", end_color="2C5F7C", fill_type="solid")
    body_font = Font(name="Arial", size=10)

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(sheet_name)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df[col_name].head(50)]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return path
